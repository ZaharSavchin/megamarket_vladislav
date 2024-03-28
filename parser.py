import time
from selenium import webdriver
from selenium.webdriver.common.by import By

from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromiumService
from openpyxl import load_workbook

from other_shops import get_other_shops


workbook = load_workbook(filename='megamarket.xlsx')
sheet = workbook.active


links_ = []
for cell in sheet['A']:
    links_.append(cell.value)

links = []


items_links = [None]
price = [None]
cashback = [None]
cashback_other = [None]
cashback_sber = [None]
delivery = [None]
pay_format = [None]
shop = [None]
attrs = [None]


with webdriver.Chrome(service=ChromiumService(ChromeDriverManager().install())) as browser:

    print(links_[1])
    # time.sleep(5)
    pages = int(links_[0])
    for i in range(1, pages + 1):
        lin = str(links_[1]).replace('/page-1', f'/page-{i}')

        print(lin)
        browser.get(lin)
        # time.sleep(3)
        lk = browser.find_elements(By.CLASS_NAME, 'catalog-item-mobile')
        print(len(lk))
        for l in lk:
            links.append(l.find_element(By.TAG_NAME, 'a').get_attribute('href'))
        # time.sleep(2)

    counter = len(links)
    print(counter)
    
    browser.maximize_window()

    for link in links:
        print(f'осталось - {counter}')
        try:
            browser.get(link)
            # time.sleep(2)

            items_links.append(link)

            try:
                browser.switch_to.frame("fl-520224")
                time.sleep(2)
                browser.find_element(By.CSS_SELECTOR, 'button.close.js-close.active').click()
                time.sleep(1)
                browser.switch_to.default_content()
                time.sleep(1)
            except Exception as err:
                print('1')

            try:
                price_ = browser.find_element(By.CLASS_NAME, 'sales-block-offer-price__price-final').text
                price_ = ''.join(price_.split(' ')[:-1])
                price.append(int(price_))
            except Exception as err:
                price_ = None
                price.append(None)

            try:
                cashback_ = browser.find_element(By.CLASS_NAME, "bonus-amount").text
                cashback_ = int(''.join(cashback_.split(' ')))
                cashback.append(cashback_)
            except Exception as err:
                cashback_ = None
                cashback.append(None)

            try:
                cashback_s = browser.find_element(By.CSS_SELECTOR,
                                                  'span.bonus-amount.bonus-amount_without-percent').text
                cashback_s = int(''.join(cashback_s.split(' '))[1:])
                cashback_sber.append(cashback_s)
            except Exception as err:
                cashback_s = None
                cashback_sber.append(None)

            try:
                cashback_o = browser.find_element(By.CSS_SELECTOR,
                                                  'div.pdp-cashback-table__money-bonus.money-bonus.xs.money-bonus_loyalty.money-bonus_grey'
                                                  ).find_element(By.CLASS_NAME, 'bonus-amount').text
                cashback_o = int(''.join(cashback_o.split(' ')))
                cashback_other.append(cashback_o)
            except Exception as err:
                cashback_o = None
                cashback_other.append(None)

            try:
                delivery_ = browser.find_element(By.CLASS_NAME, "sales-block-delivery-type__date").text
                delivery.append(delivery_)
            except Exception as err:
                delivery_ = None
                delivery.append(None)

            try:
                pay_format_ = browser.find_element(By.CLASS_NAME, "pdp-available-payment-method-block__text").text
                pay_format.append(pay_format_)
            except Exception as err:
                pay_format_ = None
                pay_format.append(None)

            try:
                shop_ = browser.find_element(By.CLASS_NAME, "pdp-merchant-rating-block__merchant-name").text
                shop.append(shop_)
            except Exception as err:
                shop_ = None
                shop.append(None)

            try:
                attr = browser.find_element(By.CLASS_NAME, "pdp-regular-attrs__characteristics").text
                attrs.append(attr)
                print(attr)
            except Exception:
                attr = None
                attrs.append(None)

            print(f'{link[:10]} = {price_}р, {cashback_}, {delivery_}, {pay_format_}, {shop_}')

            # try:
            #     get_other_shops(browser,
            #                     items_links,
            #                     price,
            #                     cashback,
            #                     cashback_other,
            #                     cashback_sber,
            #                     delivery,
            #                     pay_format,
            #                     shop
            #                     )
            # except Exception as err:
            #     print(err)
            counter -= 1
        except Exception as err:
            items_links.append(link)
            price.append(None)
            cashback.append(None)
            cashback_other.append(None)
            cashback_sber.append(None)
            delivery.append(None)
            pay_format.append(None)
            shop.append(None)
            print(f'{link} = товар не найден')
            counter -= 1

for i in range(len(items_links)):
    sheet.cell(row=i+1, column=2, value=items_links[i])

for i in range(len(price)):
    sheet.cell(row=i+1, column=3, value=price[i])

for i in range(len(cashback)):
    sheet.cell(row=i+1, column=4, value=cashback[i])

for i in range(len(cashback_other)):
    sheet.cell(row=i + 1, column=5, value=cashback_other[i])

for i in range(len(cashback_sber)):
    sheet.cell(row=i+1, column=6, value=cashback_sber[i])

for i in range(len(delivery)):
    sheet.cell(row=i+1, column=7, value=delivery[i])

for i in range(len(pay_format)):
    sheet.cell(row=i+1, column=8, value=pay_format[i])

for i in range(len(shop)):
    sheet.cell(row=i+1, column=9, value=shop[i])

for i in range(len(attrs)):
    sheet.cell(row=i+1, column=10, value=attrs[i])

workbook.save(filename='megamarket.xlsx')
